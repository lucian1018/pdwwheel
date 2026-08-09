from __future__ import annotations

import io
import re
import unicodedata
from copy import deepcopy
from datetime import date, datetime, time, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any

from cachetools import TTLCache
from openpyxl import load_workbook

from .catalog_service import normalize_name
from .config import Settings


class StockFileNotFoundError(FileNotFoundError):
    pass


def normalize_header(value: Any) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).upper()
    return re.sub(r"[^0-9A-Z\u4E00-\u9FFF]+", "", text)


MODEL_HEADERS = {
    normalize_header(value)
    for value in ("型號", "產品型號", "商品型號", "產品名稱", "商品名稱", "品名", "MODEL", "NAME")
}
SIZE_HEADERS = {
    normalize_header(value)
    for value in ("吋數", "尺寸", "輪圈尺寸", "SIZE")
}
QUANTITY_HEADERS = {
    normalize_header(value)
    for value in ("庫存", "庫存量", "數量", "台北庫存", "STOCK", "QTY", "QUANTITY")
}


def json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return str(value)


def iso_modified_time(path: Path) -> str:
    modified = datetime.fromtimestamp(path.stat().st_mtime, tz=timezone.utc)
    return modified.isoformat().replace("+00:00", "Z")


class ExcelStockService:
    def __init__(self, settings: Settings):
        self.settings = settings
        self._snapshot_cache: TTLCache = TTLCache(maxsize=2, ttl=settings.stock_cache_ttl)
        self._file_cache: TTLCache = TTLCache(maxsize=1, ttl=settings.stock_cache_ttl)

    def find_latest_file(self, force_refresh: bool = False) -> dict[str, Any]:
        if force_refresh:
            self._file_cache.clear()
        if "file" in self._file_cache:
            return self._file_cache["file"]

        inventory_dir = self.settings.inventory_dir
        inventory_dir.mkdir(parents=True, exist_ok=True)
        files = [
            path for path in inventory_dir.rglob("*")
            if path.is_file()
            and path.suffix.lower() == ".xlsx"
            and not path.name.startswith("~$")
        ]

        exact_name = self.settings.stock_excel_filename.casefold()
        prefix = self.settings.stock_excel_filename_prefix.casefold()
        if exact_name:
            candidates = [path for path in files if path.name.casefold() == exact_name]
            if not candidates:
                raise StockFileNotFoundError(
                    f"找不到指定的庫存 Excel：{self.settings.stock_excel_filename}"
                )
        elif prefix:
            candidates = [path for path in files if path.name.casefold().startswith(prefix)]
            if not candidates:
                raise StockFileNotFoundError(
                    f"找不到檔名以「{self.settings.stock_excel_filename_prefix}」開頭的 Excel"
                )
        else:
            candidates = files

        if not candidates:
            raise StockFileNotFoundError("本機庫存資料夾中找不到 Excel 庫存檔")

        candidates.sort(
            key=lambda path: (path.stat().st_mtime_ns, path.name.casefold()),
            reverse=True,
        )
        latest_path = candidates[0]
        latest = {
            "path": latest_path,
            "name": latest_path.name,
            "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "relativePath": latest_path.relative_to(self.settings.project_root).as_posix(),
            "modifiedTime": iso_modified_time(latest_path),
            "modifiedTimeNs": latest_path.stat().st_mtime_ns,
            "sizeBytes": latest_path.stat().st_size,
        }
        self._file_cache["file"] = latest
        return latest

    def get_snapshot(self, force_refresh: bool = False) -> dict[str, Any]:
        if force_refresh:
            self._snapshot_cache.clear()
        file = self.find_latest_file(force_refresh=force_refresh)
        cache_key = f"{file['relativePath']}:{file['modifiedTimeNs']}:{file['sizeBytes']}"
        if cache_key in self._snapshot_cache:
            return self._snapshot_cache[cache_key]

        snapshot = self._parse_workbook(file["path"].read_bytes(), file)
        self._snapshot_cache[cache_key] = snapshot
        return snapshot

    def _parse_workbook(self, payload: bytes, file: dict[str, Any]) -> dict[str, Any]:
        workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
        try:
            return self._parse_open_workbook(workbook, file)
        finally:
            workbook.close()

    def _parse_open_workbook(self, workbook: Any, file: dict[str, Any]) -> dict[str, Any]:
        requested_sheet = self.settings.stock_worksheet_name
        if requested_sheet:
            if requested_sheet not in workbook.sheetnames:
                raise ValueError(f"庫存 Excel 中找不到工作表：{requested_sheet}")
            worksheet = workbook[requested_sheet]
        else:
            worksheet = workbook.active

        raw_rows = worksheet.iter_rows(values_only=True)
        header_values = None
        header_row_number = 0

        for row_number, row in enumerate(raw_rows, start=1):
            values = [json_value(value) for value in row]
            if sum(value not in (None, "") for value in values) >= 2:
                header_values = values
                header_row_number = row_number
                break
            if row_number >= 30:
                break

        if not header_values:
            raise ValueError("庫存 Excel 找不到可用標題列")

        headers = self._deduplicate_headers(header_values)
        rows = []
        for row in raw_rows:
            values = [json_value(value) for value in row]
            if not any(value not in (None, "") for value in values):
                continue
            rows.append({
                header: values[index] if index < len(values) else None
                for index, header in enumerate(headers)
            })

        return {
            "file": {
                "name": file["name"],
                "mimeType": file["mimeType"],
                "relativePath": file["relativePath"],
                "modifiedTime": file["modifiedTime"],
                "sizeBytes": file["sizeBytes"],
            },
            "worksheet": worksheet.title,
            "headerRow": header_row_number,
            "headers": headers,
            "rows": rows,
            "rowCount": len(rows),
            "cacheTtlSeconds": self.settings.stock_cache_ttl,
        }

    @staticmethod
    def _deduplicate_headers(values: list[Any]) -> list[str]:
        seen: dict[str, int] = {}
        output = []
        for index, value in enumerate(values, start=1):
            base = str(value).strip() if value not in (None, "") else f"欄位{index}"
            seen[base] = seen.get(base, 0) + 1
            output.append(base if seen[base] == 1 else f"{base}_{seen[base]}")
        return output

    def merge_inventory(
        self,
        products: list[dict[str, Any]],
        snapshot: dict[str, Any],
    ) -> tuple[list[dict[str, Any]], dict[str, Any]]:
        headers = snapshot["headers"]
        model_header = self._find_header(headers, MODEL_HEADERS)
        size_header = self._find_header(headers, SIZE_HEADERS)
        quantity_header = self._find_header(headers, QUANTITY_HEADERS)

        warnings = []
        if not model_header:
            warnings.append("庫存 Excel 已取得，但找不到型號或品名欄位")
            return products, {"merged": 0, "warnings": warnings}
        if not quantity_header:
            warnings.append("庫存 Excel 已取得，但找不到庫存數量欄位")
            return products, {"merged": 0, "warnings": warnings}

        inventory_index: dict[tuple[str, int | None], list[dict[str, Any]]] = {}
        for row in snapshot["rows"]:
            model_key = normalize_name(row.get(model_header))
            if not model_key:
                continue
            size = self._extract_size(row.get(size_header)) if size_header else None
            inventory_index.setdefault((model_key, size), []).append(row)

        output = []
        merged = 0
        for source_product in products:
            product = deepcopy(source_product)
            spec = product.get("spec", {})
            product_model = normalize_name(product.get("name") or product.get("model"))
            product_size = self._extract_size(product.get("size") or spec.get("吋"))
            records = inventory_index.get((product_model, product_size))
            if records is None:
                records = inventory_index.get((product_model, None))

            if records:
                quantities = [row.get(quantity_header) for row in records]
                numeric = [float(value) for value in quantities if isinstance(value, (int, float))]
                quantity: Any = sum(numeric) if numeric else quantities[0]
                if isinstance(quantity, float) and quantity.is_integer():
                    quantity = int(quantity)
                product["inventory"] = {
                    "quantity": quantity,
                    "sourceFile": snapshot["file"]["name"],
                    "updatedAt": snapshot["file"]["modifiedTime"],
                }
                merged += 1
            else:
                product["inventory"] = None
            output.append(product)

        return output, {
            "merged": merged,
            "modelColumn": model_header,
            "sizeColumn": size_header,
            "quantityColumn": quantity_header,
            "warnings": warnings,
        }

    @staticmethod
    def _find_header(headers: list[str], aliases: set[str]) -> str | None:
        for header in headers:
            if normalize_header(header) in aliases:
                return header
        return None

    @staticmethod
    def _extract_size(value: Any) -> int | None:
        if value in (None, ""):
            return None
        match = re.search(r"(1[3-9]|2[0-2])", unicodedata.normalize("NFKC", str(value)))
        return int(match.group(1)) if match else None
